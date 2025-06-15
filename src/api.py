import asyncio, os, json, re
from TikTokApi import TikTokApi
from TikTokApi.helpers import extract_video_id_from_url
from src.export.excel import export_comment_to_excel
from openpyxl import load_workbook

def sanitize_filename(s):
    return re.sub(r'[\\/*?:"<>|]', "_", s)

def get_existing_comment_ids(filename):
    root_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    out_dir = os.path.join(root_dir, "out")
    excel_filename = os.path.join(out_dir, filename)
    if not os.path.exists(excel_filename):
        return set()
    wb = load_workbook(excel_filename)
    ws = wb.active
    return set(str(row[0].value) for row in ws.iter_rows(min_row=2) if row[0].value)

def load_state(state_file):
    if os.path.exists(state_file):
        with open(state_file, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def save_state(state_file, state):
    with open(state_file, "w", encoding="utf-8") as f:
        json.dump(state, f)

async def refresh_api(ms_token):
    api = TikTokApi()
    await api.create_sessions(headless=False, ms_tokens=[ms_token], num_sessions=1, sleep_after=3)
    return api

async def retrieve_comments(
    ms_token: str,
    video_links: list,
    target_id="",
    batch_size=1000,
    wait_time=60,
    fail_wait_time=600,
    max_failures=5,
    state_file="comment_state.json"
):
    state = load_state(state_file)
    api = await refresh_api(ms_token)
    
    if target_id:
        video_ids = [target_id]
        video_url_map = {target_id: video_links[0] if video_links else target_id}
    else:
        video_ids = [extract_video_id_from_url(video) for video in video_links]
        video_url_map = {extract_video_id_from_url(url): url for url in video_links}
    
    for current_id in video_ids:
        video_url = video_url_map.get(current_id, current_id)
        safe_filename = f"video_{sanitize_filename(video_url)}.xlsx"
        print(f"Starting to fetch comments for video {current_id}")

        # Overwrite the Excel file ONCE at the start
        export_comment_to_excel("id", "author", "text", "likes_count", safe_filename, overwrite=True)

        cursor = state.get(current_id, {}).get("cursor", 0)
        fetched_comment_ids = set()
        consecutive_failures = 0
        try:
            while True:
                try:
                    comments_gen = api.video(id=current_id).comments(count=batch_size, cursor=cursor)
                    batch = []
                    async for comment in comments_gen:
                        if comment.id in fetched_comment_ids:
                            continue
                        batch.append(comment)
                        fetched_comment_ids.add(comment.id)
                    
                    if not batch:
                        print(f"No more new comments for video {current_id}.")
                        break

                    for comment in batch:
                        export_comment_to_excel(
                            comment.id,
                            comment.author.username,
                            comment.text,
                            comment.likes_count,
                            safe_filename,
                            overwrite=False  # <-- Only overwrite ONCE at the start!
                        )
                        print(f"Fetched comment {comment.id} for video {current_id}")

                    cursor += batch_size
                    # Save state after each batch
                    state[current_id] = {"cursor": cursor}
                    save_state(state_file, state)
                    print(f"Waiting {wait_time} seconds before fetching next batch for video {current_id}...")
                    await asyncio.sleep(wait_time)
                    consecutive_failures = 0 
                except Exception as e:
                    err_msg = str(e).lower()
                    print(f"Error while fetching comments for video {current_id}: {e}")
                    if "bot" in err_msg or "fail" in err_msg or "empty response" in err_msg:
                        consecutive_failures += 1
                        print(f"Rate limited or detected as bot. Failure count: {consecutive_failures}/{max_failures}.")
                        print(f"⚠️  Please open this video in your browser to get a fresh MS_TOKEN:\n{video_url}\n")
                        # Prompt for new MS_TOKEN
                        new_token = input("You have been rate limited or your MS_TOKEN expired. Please enter a new MS_TOKEN (or press Enter to reuse the old one): ").strip()
                        if new_token:
                            ms_token = new_token
                            api = await refresh_api(ms_token)
                            print("Refreshed session with new MS_TOKEN. Waiting 60 seconds before retrying...")
                            await asyncio.sleep(60)
                        else:
                            print(f"Waiting {fail_wait_time // 60} minutes before retrying with the same token...")
                            await asyncio.sleep(fail_wait_time)
                        state[current_id] = {"cursor": cursor}
                        save_state(state_file, state)
                        if consecutive_failures >= max_failures:
                            print(f"Flagged {max_failures} times in a row for video {current_id}. Saving progress and stopping for this video.")
                            state[current_id] = {"cursor": cursor}
                            save_state(state_file, state)
                            break
                    else:
                        # Save state on unknown error and break
                        state[current_id] = {"cursor": cursor}
                        save_state(state_file, state)
                        break
        except KeyboardInterrupt:
            print("Interrupted by user. Saving state and exiting...")
            state[current_id] = {"cursor": cursor}
            save_state(state_file, state)
            break

    if os.path.exists(state_file):
        try:
            os.remove(state_file)
            print(f"Deleted state file: {state_file}")
        except Exception as e:
            print(f"Could not delete state file: {e}")

    print("✅ All videos processed. Shutting down.")

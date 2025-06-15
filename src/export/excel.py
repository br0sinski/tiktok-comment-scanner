import os
from openpyxl import Workbook, load_workbook

def export_comment_to_excel(id: str, author: str, text: str, likes_count: int, filename: str, overwrite=False):
    root_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    out_dir = os.path.join(root_dir, "out")
    os.makedirs(out_dir, exist_ok=True)

    excel_filename = os.path.join(out_dir, filename)

    if overwrite and os.path.exists(excel_filename):
        os.remove(excel_filename)

    file_exists = os.path.exists(excel_filename)

    if file_exists:
        wb = load_workbook(excel_filename)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["id", "author", "text", "likes_count"])

    ws.append([id, author, text, likes_count])
    wb.save(excel_filename)